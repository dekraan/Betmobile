import json
from datetime import datetime

import pandas as pd
import numpy as np
from psycopg2.extras import execute_values
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from config import OUTPUT_DIR, ECI_RULE_PARAMS, USE_CUTOFF_FEATURES, MIN_STRENGTH
from db import db_conn
from utils import _json_default, safe_float, choose_relevant_side



# =====================================================================
# SAVE TO DB (public.picks_evaluated)
# =====================================================================
def save_to_db(picks: pd.DataFrame, calib_meta: dict | None = None):

    cfg = dict(
        version="eci_picks v4.1 calibrated",
        generated_at=datetime.now().isoformat(),
        params=ECI_RULE_PARAMS,
        use_cutoff=USE_CUTOFF_FEATURES,
        calibration=calib_meta or {},
        min_strength=MIN_STRENGTH,
    )

    with db_conn() as conn, conn.cursor() as cur:
        cur.execute(
            """
            INSERT INTO public.picks_run (created_by, config, use_cutoff, source_note)
            VALUES (current_user, %s::jsonb, %s, 'eci v4.1 calibrated')
            RETURNING run_id;
            """,
            (json.dumps(cfg, default=_json_default), USE_CUTOFF_FEATURES),
        )
        run_id = cur.fetchone()[0]

        if picks.empty:
            conn.commit()
            print("Geen picks om op te slaan.")
            return run_id

        rows = []
        for _, r in picks.iterrows():
            # Edge = EV - 1
            edge_h = float(r["bet_home"] - 1.0) if not pd.isna(r["bet_home"]) else None
            edge_d = float(r["bet_draw"] - 1.0) if not pd.isna(r["bet_draw"]) else None
            edge_a = float(r["bet_away"] - 1.0) if not pd.isna(r["bet_away"]) else None

            # legacy rule_a/b/c mappen we 1-op-1 op AwayRule/HomeRule
            rule_a = bool(r["AwayRule"])
            rule_b = bool(r["HomeRule"])
            rule_c = False

            # pick_reason: gebruik Advice als die bestaat, anders Selection
            pick_reason = r.get("Advice") or r.get("Selection") or "ECI"

            if r.get("Selection") == "HOME":
                selected_raw_strength = r.get("RawStrength_Home", r.get("RuleStrength"))
                selected_adj_strength = r.get("RuleStrengthAdj_Home", r.get("RuleStrengthAdj"))
                selected_reason = r.get("Home_reason", r.get("Rule_reason"))
            elif r.get("Selection") == "AWAY":
                selected_raw_strength = r.get("RawStrength_Away", r.get("RuleStrength"))
                selected_adj_strength = r.get("RuleStrengthAdj_Away", r.get("RuleStrengthAdj"))
                selected_reason = r.get("Away_reason", r.get("Rule_reason"))
            else:
                selected_raw_strength = r.get("RuleStrength")
                selected_adj_strength = r.get("RuleStrengthAdj")
                selected_reason = r.get("Rule_reason")

            rule_strength = float(selected_raw_strength) if pd.notna(selected_raw_strength) else None
            rule_strength_adj = float(selected_adj_strength) if pd.notna(selected_adj_strength) else None
            pick_type = r.get("PickType")

            hours_stale = r.get("hours_stale")
            stake_rec = None
            score = r.get("score")
            result = None
            outcome = None
            settled_at = None
            date_ts = None

            rule_passed = bool(r.get("rule_passed", True))
            rule_reason = selected_reason

            rows.append(
                (
                    run_id,
                    r["match_id"],
                    r["competition"],
                    r["date"],
                    r["home_team"],
                    r["away_team"],
                    r["odds_home"],
                    r["odds_draw"],
                    r["odds_away"],
                    r["Home Prob"],
                    r["Draw Prob"],
                    r["Away Prob"],
                    r["bet_home"],
                    r["bet_draw"],
                    r["bet_away"],
                    edge_h,
                    edge_d,
                    edge_a,
                    rule_a,
                    rule_b,
                    rule_c,
                    pick_reason,
                    rule_strength,
                    rule_strength_adj,
                    r.get("away_drift_pct"),
                    r.get("home_drift_pct"),
                    r.get("n_snapshots"),
                    hours_stale,
                    r["Selection"],
                    stake_rec,
                    score,
                    result,
                    outcome,
                    settled_at,
                    date_ts,
                    rule_passed,
                    rule_reason,
                    pick_type,
                    safe_float(r.get("value_edge")),
                    safe_float(r.get("prob_edge")),
                    safe_float(r.get("drift_score")),
                    int(r.get("snapshot_count")) if pd.notna(r.get("snapshot_count")) else None,
                    r.get("strength_bucket"),
                    safe_float(r.get("rating_gap")),
                    safe_float(r.get("drift_range")),
                    r.get("pick_tier"),
                    int(r.get("pick_stars")) if pd.notna(r.get("pick_stars")) else None,
                    r.get("sector_tags"),
                    r.get("danger_tags"),
                    r.get("classification_reason"),
                    bool(r.get("passes_danger_combo_v2")) if pd.notna(r.get("passes_danger_combo_v2")) else None,
                    bool(r.get("passes_danger_combo_v2_no_longshots")) if pd.notna(r.get("passes_danger_combo_v2_no_longshots")) else None,
                )
            )

        sql = """
            INSERT INTO public.picks_evaluated (
                run_id, match_id, competition, date,
                home_team, away_team,
                odds_home, odds_draw, odds_away,
                prob_home, prob_draw, prob_away,
                bet_home, bet_draw, bet_away,
                edge_h, edge_d, edge_a,
                rule_a, rule_b, rule_c,
                pick_reason, rule_strength, rule_strength_adj,
                away_drift_pct, home_drift_pct, n_snapshots,
                hours_stale, selection, stake_rec,
                score, result, outcome, settled_at, date_ts,
                rule_passed, rule_reason, pick_type,
                value_edge, prob_edge, drift_score, snapshot_count,
                strength_bucket, rating_gap, drift_range,
                pick_tier, pick_stars, sector_tags, danger_tags, classification_reason,
                passes_danger_combo_v2, passes_danger_combo_v2_no_longshots
            ) VALUES %s
        """

        execute_values(cur, sql, rows, page_size=500)
        conn.commit()

    print(f"Picks saved: {len(rows)}")
    return run_id

def save_single_fails_to_db(run_id: int, today_view: pd.DataFrame):
    if today_view is None or today_view.empty:
        print("Geen today_view om single-fails op te slaan.")
        return

    sf = today_view[today_view["Bucket"] == "SINGLE_FAIL"].copy()

    if sf.empty:
        print("Geen single-fail candidates om op te slaan.")
        return

    rows = []

    for _, r in sf.iterrows():
        side = r.get("TodaySide")

        if side == "HOME":
            odds = r.get("odds_home")
            probability = r.get("Home Prob")
            value_score = r.get("bet_home")
            drift_pct = r.get("home_drift_pct")
        elif side == "AWAY":
            odds = r.get("odds_away")
            probability = r.get("Away Prob")
            value_score = r.get("bet_away")
            drift_pct = r.get("away_drift_pct")
        else:
            odds = None
            probability = None
            value_score = None
            drift_pct = None

        rows.append((
            run_id,
            r.get("match_id"),
            r.get("date"),
            r.get("competition"),
            r.get("home_team"),
            r.get("away_team"),

            side,
            r.get("TodayReason"),
            safe_float(r.get("single_fail_margin")),
            safe_float(r.get("snap_needed")),

            safe_float(r.get("SingleFailRawStrength")),
            safe_float(r.get("SingleFailAdjStrength")),
            safe_float(r.get("SingleFailCalibratedStrength")),

            safe_float(odds),
            safe_float(probability),
            safe_float(value_score),

            safe_float(r.get("odds_home")),
            safe_float(r.get("odds_draw")),
            safe_float(r.get("odds_away")),
            safe_float(r.get("Home Prob")),
            safe_float(r.get("Draw Prob")),
            safe_float(r.get("Away Prob")),
            safe_float(r.get("bet_home")),
            safe_float(r.get("bet_draw")),
            safe_float(r.get("bet_away")),

            safe_float(r.get("rating_gap")),
            safe_float(r.get("rating_home_edge")),
            int(r.get("n_snapshots")) if pd.notna(r.get("n_snapshots")) else None,
            safe_float(drift_pct),
            safe_float(r.get("home_drift_pct")),
            safe_float(r.get("away_drift_pct")),

            r.get("home_fail_reasons"),
            r.get("away_fail_reasons"),
            int(r.get("home_fail_count")) if pd.notna(r.get("home_fail_count")) else None,
            int(r.get("away_fail_count")) if pd.notna(r.get("away_fail_count")) else None,

            r.get("score"),
            None,
            None,
            None,
        ))

    sql = """
        INSERT INTO public.picks_single_fail_candidates (
            run_id,
            match_id, date, competition, home_team, away_team,
            side, fail_reason, single_fail_margin, snap_needed,
            single_fail_raw_strength,
            single_fail_adj_strength,
            single_fail_calibrated_strength,
            odds, probability, value_score,
            odds_home, odds_draw, odds_away,
            prob_home, prob_draw, prob_away,
            bet_home, bet_draw, bet_away,
            rating_gap, rating_home_edge,
            n_snapshots, drift_pct, home_drift_pct, away_drift_pct,
            home_fail_reasons, away_fail_reasons,
            home_fail_count, away_fail_count,
            score, result, outcome, settled_at
        ) VALUES %s
    """

    with db_conn() as conn, conn.cursor() as cur:
        execute_values(cur, sql, rows, page_size=500)
        conn.commit()

    print(f"Single-fail candidates saved: {len(rows)}")

# =====================================================================
# EXCEL EXPORT
# =====================================================================
def save_excel(picks: pd.DataFrame, df_all: pd.DataFrame, snap_watch: pd.DataFrame, price_watch: pd.DataFrame, today_view: pd.DataFrame, near_miss: pd.DataFrame | None = None):
    out = OUTPUT_DIR / f"eci_picks_{datetime.now().strftime('%Y%m%d')}.xlsx"

    def _compact(df: pd.DataFrame, extra_cols=None):
        """Houd sheets leesbaar: kernkolommen + optionele extras."""
        if df is None or df.empty:
            return df
        base_cols = [
            "match_id","date","competition","home_team","away_team",
            "odds_home","odds_draw","odds_away",
            "Home Prob","Draw Prob","Away Prob",
            "bet_home","bet_draw","bet_away",
            "home_rating","away_rating","rating_gap","rating_home_edge",
            "n_snapshots",
            "home_drift_pct","away_drift_pct",
            "home_drift_abs","away_drift_abs",
            "home_last_move_pct","away_last_move_pct",
            "home_recent24_pct","away_recent24_pct",
            "home_range","away_range",
            "hours_stale","market_age_hours",
            "hours_to_kickoff","scrape_to_kickoff_hours",
            "RuleStrength","RuleStrengthAdj","RuleStrengthCalibrated",
            "DriftNotes",
        ]
        # only keep existing
        cols = [c for c in base_cols if c in df.columns]
        if extra_cols:
            cols += [c for c in extra_cols if c in df.columns and c not in cols]
        return df[cols].copy()

    def _build_summary(df_all: pd.DataFrame):
        """Samenvatting van single-fail redenen per competitie + totaal, op basis van de relevante kant."""
        if df_all is None or df_all.empty:
            return pd.DataFrame({"info": ["Geen data"]})

        df = df_all.copy()

        df["single_fail_side"] = None
        df["single_fail_reason"] = None

        df["single_fail_side"] = df.apply(
            lambda r: choose_relevant_side(
                r,
                home_condition=(r.get("home_fail_count") == 1),
                away_condition=(r.get("away_fail_count") == 1),
            ),
            axis=1
        )

        df["single_fail_reason"] = np.where(
            df["single_fail_side"] == "HOME",
            df.get("home_fail_reasons"),
            np.where(
                df["single_fail_side"] == "AWAY",
                df.get("away_fail_reasons"),
                None
            )
        )

        # Alleen gevallen met een gekozen relevante kant
        sf = df[df["single_fail_side"].notna()].copy()
        if sf.empty:
            return pd.DataFrame({"info": ["Geen single-fail gevallen"]})

        # Totaal
        total = (
            sf.groupby(["single_fail_reason"], observed=True)
              .size()
              .reset_index(name="count")
              .sort_values("count", ascending=False)
        )
        total.insert(0, "competition", "ALL")

        # Per competitie
        per_comp = (
            sf.groupby(["competition", "single_fail_reason"], observed=True)
              .size()
              .reset_index(name="count")
              .sort_values(["competition", "count"], ascending=[True, False])
        )

        return pd.concat([total, per_comp], ignore_index=True)

    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        # PICKS
        if picks.empty:
            pd.DataFrame({"info": ["Geen picks"]}).to_excel(writer, sheet_name="PICKS", index=False)
        else:
            _compact(
                picks,
                extra_cols=[
                    "pick_tier", "pick_stars", "sector_tags", "danger_tags",
                    "classification_reason", "passes_danger_combo_v2",
                    "passes_danger_combo_v2_no_longshots",
                    "PickType", "Selection", "Advice",
                    "SelectedRawStrength", "SelectedAdjStrength", "SelectedStrength",
                    "SecondaryStrength",
                    "selected_odds", "selected_prob", "selected_value_score",
                    "odds_bucket", "prob_bucket", "rating_gap_bucket", "value_bucket", "market_support",
                    "DriftNotes", "DriftNotes_Home", "DriftNotes_Away",
                    "Rule_reason"
                ]
            ).to_excel(writer, sheet_name="PICKS", index=False)

        # ALL_MATCHES (laat je desnoods full; ik maak 'm compact zodat het werkbaar blijft)
        _compact(df_all).to_excel(writer, sheet_name="ALL_MATCHES", index=False)

        # TODAY (compact): 2 blokken onder elkaar (PICKS + SINGLE_FAIL)
        if today_view is None or today_view.empty:
            pd.DataFrame({"info": ["Geen wedstrijden voor vandaag"]}).to_excel(
                writer, sheet_name="TODAY", index=False
            )
        else:
            tv = today_view.copy()

            strength_col = "RuleStrengthCalibrated" if "RuleStrengthCalibrated" in tv.columns else "RuleStrengthAdj"

            # compacte kolommen
            base_today_cols = [
                "Bucket",
                "match_id","date","competition","home_team","away_team",
                "PickType","Selection","Advice",
                "pick_tier", "pick_stars", "sector_tags", "danger_tags",
                "classification_reason", "passes_danger_combo_v2",
                "passes_danger_combo_v2_no_longshots",
                "TodaySide","TodayReason",
                "single_fail_margin","snap_needed",
                "SingleFailRawStrength",
                "SingleFailAdjStrength",
                "SingleFailCalibratedStrength",
                "odds_home","odds_draw","odds_away",
                "Home Prob","Draw Prob","Away Prob",
                "bet_home","bet_draw","bet_away",
                "rating_gap","rating_home_edge",
                "n_snapshots","home_drift_pct","away_drift_pct",
                "home_fail_reasons","away_fail_reasons","home_fail_count","away_fail_count",
                strength_col,
            ]
            cols = [c for c in base_today_cols if c in tv.columns]
            tv = tv[cols].copy()

            # split
            picks_part = tv[tv["Bucket"] == "PICK"].copy()
            sf_part    = tv[tv["Bucket"] == "SINGLE_FAIL"].copy()

            # single-fails groeperen op reason (netter)
            if not sf_part.empty and "TodayReason" in sf_part.columns:
                sf_part = sf_part.sort_values(["TodayReason", "single_fail_margin", strength_col], ascending=[True, False, False])

            ws_name = "TODAY"

            row = 0
            if picks_part.empty:
                pd.DataFrame({"info": ["TODAY: geen picks"]}).to_excel(
                    writer, sheet_name=ws_name, index=False, startrow=row
                )
                row += 2
            else:
                pd.DataFrame({"section": ["TODAY PICKS"]}).to_excel(
                    writer, sheet_name=ws_name, index=False, startrow=row
                )
                row += 2
                picks_part.to_excel(writer, sheet_name=ws_name, index=False, startrow=row)
                row += len(picks_part) + 3

            if sf_part.empty:
                pd.DataFrame({"info": ["TODAY: geen single-fail kandidaten"]}).to_excel(
                    writer, sheet_name=ws_name, index=False, startrow=row
                )
            else:
                pd.DataFrame({"section": ["TODAY SINGLE_FAIL (per reason gegroepeerd)"]}).to_excel(
                    writer, sheet_name=ws_name, index=False, startrow=row
                )
                row += 2
                sf_part.to_excel(writer, sheet_name=ws_name, index=False, startrow=row)


        # WATCH_SNAP (compact + extras)
        if snap_watch.empty:
            pd.DataFrame({"info": ["Geen snap-watch items"]}).to_excel(writer, sheet_name="WATCH_SNAP", index=False)
        else:
            _compact(
                snap_watch,
                extra_cols=["WatchSide","home_fail_reasons","away_fail_reasons","home_fail_count","away_fail_count",
                            "snap_needed","closest_side","closest_fail_reasons"]
            ).to_excel(writer, sheet_name="WATCH_SNAP", index=False)

        # WATCH_PRICE (compact + extras)
        if price_watch.empty:
            pd.DataFrame({"info": ["Geen price-watch items"]}).to_excel(writer, sheet_name="WATCH_PRICE", index=False)
        else:
            _compact(
                price_watch,
                extra_cols=["WatchSide","WatchReason","value_margin","odds_margin",
                            "home_fail_reasons","away_fail_reasons","home_fail_count","away_fail_count",
                            "closest_side","closest_fail_reasons"]
            ).to_excel(writer, sheet_name="WATCH_PRICE", index=False)

        # NEAR_MISS
        if near_miss is None or near_miss.empty:
            pd.DataFrame({"info": ["Geen near-miss items"]}).to_excel(
                writer, sheet_name="NEAR_MISS", index=False
            )
        else:
            _compact(
                near_miss,
                extra_cols=[
                    "NearMissRank",
                    "single_fail_margin",
                    "NearMissSide", "NearMissReason",
                    "prob_margin", "value_margin", "odds_margin",
                    "drift_margin", "rating_margin", "edge_margin",
                    "snap_needed",
                    "selected_prob_sf", "selected_value_sf",
                    "selected_odds_sf", "selected_drift_sf",
                    "home_fail_reasons", "away_fail_reasons",
                    "home_fail_count", "away_fail_count",
                    "RuleStrengthCalibrated", "RuleStrengthAdj",
                    "closest_side", "closest_fail_reasons",
                ],
            ).to_excel(writer, sheet_name="NEAR_MISS", index=False)

        # WATCH_SUMMARY
        summary = _build_summary(df_all)
        summary.to_excel(writer, sheet_name="WATCH_SUMMARY", index=False)

        # Extra per competitie (optioneel: ook compact)
        for comp, df_comp in df_all.groupby("competition"):
            safe = comp.replace("/", "_").replace("\\", "_")[:28]
            _compact(df_comp).to_excel(writer, sheet_name=safe, index=False)

    # Kleur de picks-sheet groen
    if picks.empty:
        print("Excel gemaakt (geen picks).")
        return

    wb = load_workbook(out)
    ws = wb["PICKS"]
    fillPick = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    for r in range(2, ws.max_row + 1):
        for c in ws[r]:
            c.fill = fillPick

    wb.save(out)
    print(f"Excel gemaakt: {out}")